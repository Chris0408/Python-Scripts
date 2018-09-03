import pymysql
from openpyxl import Workbook

#  连接数据库信息
conn = pymysql.Connect(
    host='192.168.0.122',
    port=3306,
    user='root',
    passwd='123456',
    db='test',
    charset='utf8')
cur = conn.cursor()

sql = 'select * from vpn'

# 执行查询
cur.execute(sql)

# 保存结果
result = cur.fetchall()

# 关闭连接
cur.close
conn.close

# 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
wb = Workbook()

# 获取当前活跃的worksheet,默认就是第一个worksheet
ws = wb.active

# 设置Excel头
ws.title = "VPN开通名单"

# 设置Excel第一行的标签名字
# 这里有个小bug pymysql读出的数据，字段排序是乱的。可能会与下边的列对不上。有待后续修复。
# 目前最简单的方法是先输出一次result，看下顺序，在确定下边的列名：）
ws.cell(row=1, column=1).value = "序号"
ws.cell(row=1, column=2).value = "姓名"
ws.cell(row=1, column=3).value = "用户名"
ws.cell(row=1, column=4).value = "密码"
ws.cell(row=1, column=5).value = "开通日期"
ws.cell(row=1, column=6).value = "备注"

# 添加从数据表中查出来的数据
for i in result:
    print(i)
    ws.append(i)

# 保存到本地Excel
excel_file_name = "/Users/user/Desktop/t.xlsx"
wb.save(excel_file_name)
print("导入完成")
wb.close